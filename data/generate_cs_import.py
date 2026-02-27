#!/usr/bin/env python3
"""
Generate CSV import files for C&S 2026 Charge Maps from Excel.

Creates 3 CSV files to be imported sequentially:
1. opportunities.csv - One Opportunity per Account (Renewal record type)
2. charge_maps.csv - One ChargeMap per Account (linked to Opportunity)
3. charge_map_items.csv - One ChargeMapItem per line in the Excel
"""

import csv
import openpyxl
from collections import OrderedDict
from datetime import datetime

EXCEL_PATH = "/Users/clementboschet/Documents/repos/pixid-salesforce/PIXID - création ChargeMap C&S.xlsx"
SHEET_NAME = "DOCUMENT DE TRAVAIL"
OUTPUT_DIR = "/Users/clementboschet/Documents/repos/pixid-salesforce/data/import_cs_2026"

# Salesforce constants
OPP_RECORD_TYPE_RENEWAL = "012IV000001jKlBYAU"
CHARGEMAP_RECORD_TYPE_PIXID_FR = "012IV000001yk7WYAQ"

# Field mappings from Excel values to Salesforce picklist values
ACTIVITY_MAP = {
    "Pixid France": "Pixid FR",
}

PAYMENT_METHOD_MAP = {
    "Transfer": "Transfer",
    "Direct debit": "Direct Debit",
}

BILLING_CONDITIONS_MAP = {
    30: "30 days",
    45: "45 days",
    60: "60 days",
}

PO_REQUIRED_MAP = {
    "True": "true",
    "False": "false",
    True: "true",
    False: "false",
}


def parse_date(val):
    """Parse a date value from Excel (could be datetime or string)."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        # Try common formats
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
    return str(val)


def safe_str(val):
    """Convert value to string, returning empty string for None."""
    if val is None:
        return ""
    return str(val).strip()


def main():
    import os
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    # Read all data rows (starting from row 8, header is row 7)
    rows = []
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, values_only=True):
        ext_id = row[3]  # External ID (Account Id)
        if ext_id:
            rows.append(row)

    print(f"Total data rows: {len(rows)}")

    # =========================================================================
    # 1. Group by Account (External ID) for Opportunities and ChargeMaps
    # =========================================================================
    accounts = OrderedDict()
    for row in rows:
        ext_id = safe_str(row[3])
        if ext_id not in accounts:
            accounts[ext_id] = {
                "account_id": ext_id,
                "account_name": safe_str(row[2]),  # Nom SF
                "account_number": safe_str(row[0]),  # N° Compte
                "activity": safe_str(row[11]),
                "offer": safe_str(row[12]),
                "closing_date": parse_date(row[13]),
                "legacy_id": safe_str(row[14]),
                # ChargeMap fields (take from first row of this account)
                "invoice_sending_method": safe_str(row[17]),
                "send_invoice_to": safe_str(row[18]),
                "payment_method": safe_str(row[19]),
                "charge_map_type": safe_str(row[20]),
                "billing_business_name": safe_str(row[21]),
                "billing_department": safe_str(row[22]),
                "billing_street": safe_str(row[23]),
                "billing_city": safe_str(row[24]),
                "billing_postal_code": safe_str(row[25]),
                "billing_instructions": safe_str(row[26]),
                "billing_conditions": row[27],
                "po_required": safe_str(row[28]),
                "po_reference": safe_str(row[29]),
                "po_request_date": parse_date(row[30]),
                "shipping_business_name": safe_str(row[31]),
                "shipping_department": safe_str(row[32]),
                "shipping_street": safe_str(row[33]),
                "shipping_city": safe_str(row[34]),
                "billing_contact_id": safe_str(row[35]),
                "cash_collection_id": safe_str(row[37]),
                "items": [],
            }
        # Add item for this row
        accounts[ext_id]["items"].append(row)

    print(f"Unique accounts: {len(accounts)}")

    # =========================================================================
    # 2. Generate Opportunities CSV
    # =========================================================================
    opp_file = os.path.join(OUTPUT_DIR, "opportunities.csv")
    opp_fields = [
        "AccountId",
        "RecordTypeId",
        "Name",
        "StageName",
        "CloseDate",
        "Activity__c",
        "Offer__c",
        "CurrencyIsoCode",
        "ChannelType__c",
    ]

    with open(opp_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=opp_fields)
        writer.writeheader()

        for ext_id, acct in accounts.items():
            activity = ACTIVITY_MAP.get(acct["activity"], acct["activity"])
            opp_name = f"{acct['account_name']} - C&S - 2026"

            writer.writerow({
                "AccountId": ext_id,
                "RecordTypeId": OPP_RECORD_TYPE_RENEWAL,
                "Name": opp_name,
                "StageName": "Closed - Won",
                "CloseDate": "2026-01-01",
                "Activity__c": activity,
                "Offer__c": acct["offer"],
                "CurrencyIsoCode": "EUR",
                "ChannelType__c": "Direct",
            })

    print(f"Written {len(accounts)} opportunities to {opp_file}")

    # =========================================================================
    # 3. Generate Charge Maps CSV
    #    (SourceOpportunity__c will be filled after Opportunity import)
    # =========================================================================
    cm_file = os.path.join(OUTPUT_DIR, "charge_maps.csv")
    cm_fields = [
        "Account__c",
        "RecordTypeId",
        "Type__c",
        "Status__c",
        "CurrencyIsoCode",
        "AutoRenewal__c",
        "Primary__c",
        "BillingConditions__c",
        "InvoiceSendingMethod__c",
        "PaymentMethod__c",
        "SendInvoiceTo__c",
        "BillingInstructions__c",
        "PurchaseOrderRequired__c",
        "PurchaseOrderReference__c",
        "PurchaseOrderRequestDate__c",
        "BillingBusinessName__c",
        "BillingDepartment__c",
        "BillingAddress__Street__s",
        "BillingAddress__City__s",
        "BillingAddress__PostalCode__s",
        "BillingAddress__CountryCode__s",
        "ShippingBusinessName__c",
        "ShippingDepartment__c",
        "ShippingAddress__Street__s",
        "ShippingAddress__City__s",
        "ShippingAddress__CountryCode__s",
        "BillingContact__c",
        "CashCollectionContact__c",
        "SourceOpportunity__c",
    ]

    with open(cm_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cm_fields)
        writer.writeheader()

        for ext_id, acct in accounts.items():
            billing_cond = BILLING_CONDITIONS_MAP.get(acct["billing_conditions"], "")
            payment = PAYMENT_METHOD_MAP.get(acct["payment_method"], acct["payment_method"])
            po_req = PO_REQUIRED_MAP.get(acct["po_required"], "false")

            writer.writerow({
                "Account__c": ext_id,
                "RecordTypeId": CHARGEMAP_RECORD_TYPE_PIXID_FR,
                "Type__c": "Renewal",
                "Status__c": "Backlog",
                "CurrencyIsoCode": "EUR",
                "AutoRenewal__c": "true",
                "Primary__c": "false",
                "BillingConditions__c": billing_cond,
                "InvoiceSendingMethod__c": acct["invoice_sending_method"],
                "PaymentMethod__c": payment if payment else "",
                "SendInvoiceTo__c": acct["send_invoice_to"],
                "BillingInstructions__c": acct["billing_instructions"],
                "PurchaseOrderRequired__c": po_req,
                "PurchaseOrderReference__c": acct["po_reference"] if acct["po_reference"] != "YES" else "",
                "PurchaseOrderRequestDate__c": acct["po_request_date"],
                "BillingBusinessName__c": acct["billing_business_name"],
                "BillingDepartment__c": acct["billing_department"],
                "BillingAddress__Street__s": acct["billing_street"],
                "BillingAddress__City__s": acct["billing_city"],
                "BillingAddress__PostalCode__s": acct["billing_postal_code"],
                "BillingAddress__CountryCode__s": "FR",
                "ShippingBusinessName__c": acct["shipping_business_name"],
                "ShippingDepartment__c": acct["shipping_department"],
                "ShippingAddress__Street__s": acct["shipping_street"],
                "ShippingAddress__City__s": acct["shipping_city"],
                "ShippingAddress__CountryCode__s": "FR",
                "BillingContact__c": acct["billing_contact_id"],
                "CashCollectionContact__c": acct["cash_collection_id"],
                "SourceOpportunity__c": "",  # To be filled after Opportunity import
            })

    print(f"Written {len(accounts)} charge maps to {cm_file}")

    # =========================================================================
    # 4. Generate Charge Map Items CSV
    #    (ChargeMap__c will be filled after ChargeMap import)
    # =========================================================================
    cmi_file = os.path.join(OUTPUT_DIR, "charge_map_items.csv")
    cmi_fields = [
        "AccountId_ref",  # Helper column to link to ChargeMap (not a SF field)
        "ChargeMap__c",
        "Product__c",
        "Name",
        "LineDescription__c",
        "BillingSchedule__c",
        "Term__c",
        "Quantity__c",
        "Sales_unit_price__c",
        "Amount__c",
        "Status__c",
        "StartDateRevRec__c",
        "End_date__c",
        "CurrencyIsoCode",
        "PO_Required__c",
        "Activated__c",
    ]

    item_count = 0
    with open(cmi_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cmi_fields)
        writer.writeheader()

        for ext_id, acct in accounts.items():
            for item_row in acct["items"]:
                quantity = item_row[39] if item_row[39] is not None else 1
                unit_price = item_row[40] if item_row[40] is not None else 0
                amount = item_row[41] if item_row[41] is not None else 0

                writer.writerow({
                    "AccountId_ref": ext_id,
                    "ChargeMap__c": "",  # To be filled after ChargeMap import
                    "Product__c": safe_str(item_row[4]),  # Products ID
                    "Name": safe_str(item_row[5]),  # Article SF
                    "LineDescription__c": safe_str(item_row[6]),
                    "BillingSchedule__c": safe_str(item_row[7]),
                    "Term__c": safe_str(item_row[8]),
                    "Quantity__c": quantity,
                    "Sales_unit_price__c": unit_price,
                    "Amount__c": amount,
                    "Status__c": "Backlog",
                    "StartDateRevRec__c": parse_date(item_row[9]),
                    "End_date__c": parse_date(item_row[10]),
                    "CurrencyIsoCode": "EUR",
                    "PO_Required__c": "false",
                    "Activated__c": "false",
                })
                item_count += 1

    print(f"Written {item_count} charge map items to {cmi_file}")
    print(f"\nAll files generated in: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
